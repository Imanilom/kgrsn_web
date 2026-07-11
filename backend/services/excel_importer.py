"""
Excel Importer — membaca Rekap Mei-Juni 2026.xlsx dan menyimpan histori harga
ke tabel price_history.
"""
import re
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

import models

logger = logging.getLogger(__name__)

# ── Konstanta Sheet ──────────────────────────────────────────────────────────
INVOICE_BB_SHEETS = [
    "Invoice B01", "Invoice B02", "Invoice Wer",
    "Invoice Bin", "Invoice Teg", "Invoice Kal",
]

DAPUR_MAP = {
    "Invoice B01": "Bakung 01",
    "Invoice B02": "Bakung 02",
    "Invoice Wer": "Werkudara",
    "Invoice Bin": "Binawan",
    "Invoice Teg": "Tegalsari",
    "Invoice Kal": "Kalikoa",
}

# ── Normalisasi nama barang ──────────────────────────────────────────────────
ALIAS_MAP = {
    "cabe merah": "cabai merah",
    "cabe rawit": "cabai rawit",
    "cabe hijau": "cabai hijau",
    "bwg merah": "bawang merah",
    "bwg putih": "bawang putih",
    "minyak goreng": "minyak",
    "minyak kg": "minyak",
    "minyak 2lt": "minyak",
    "minyak ltr": "minyak",
    "ayam broiler": "ayam potong",
    "ayam buras": "ayam kampung",
    "sgm": "susu",
    "knorr sapi": "knorr",
    "masako": "penyedap",
    "royco": "penyedap",
    "ajinomoto": "penyedap",
    "gula pasir lokal": "gula pasir",
    "gula putih": "gula pasir",
}


def normalize_nama(nama: str) -> str:
    """Normalisasi nama barang ke format standar."""
    if not nama:
        return ""
    nama = str(nama).strip().lower()
    # Hapus karakter khusus
    nama = re.sub(r"[^\w\s]", " ", nama)
    nama = re.sub(r"\s+", " ", nama).strip()
    # Apply alias
    return ALIAS_MAP.get(nama, nama)


def parse_rupiah(val) -> Optional[float]:
    """Parse nilai rupiah dari Excel (bisa berupa string '12.000' atau float 12000)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = float(val)
        return v if v > 0 else None
    # String
    s = str(val).strip().replace(".", "").replace(",", "").replace("Rp", "").strip()
    try:
        v = float(s)
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def parse_tanggal_blok(cell_value) -> Optional[date]:
    """
    Parse baris tanggal blok di invoice, misal:
    '📅  Sunday, 24 May 2026' atau 'Sunday, 24 May 2026'
    """
    if not cell_value:
        return None
    s = str(cell_value).strip()
    # Hapus emoji & whitespace berlebih
    s = re.sub(r"[^\w\s,]", "", s).strip()
    # Coba format: "Sunday 24 May 2026" atau "24 May 2026"
    patterns = [
        "%A %d %B %Y",
        "%A, %d %B %Y",
        "%d %B %Y",
        "%d/%m/%Y",
    ]
    for pat in patterns:
        try:
            return datetime.strptime(s, pat).date()
        except ValueError:
            pass
    return None


def is_date_row(row) -> bool:
    """Cek apakah baris ini adalah baris header tanggal (bukan item)."""
    first = str(row[0] or "").strip() if row else ""
    # Baris tanggal biasanya dimulai dengan ikon 📅 atau nama hari
    days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    return any(d in first.lower() for d in days) or "📅" in first


def remove_outliers_iqr(prices: list[float]) -> list[float]:
    """Hapus outlier menggunakan metode IQR."""
    if len(prices) < 4:
        return prices
    sorted_p = sorted(prices)
    n = len(sorted_p)
    q1 = sorted_p[n // 4]
    q3 = sorted_p[(3 * n) // 4]
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    return [p for p in prices if lower <= p <= upper]


def import_excel_to_db(filepath: str, db: Session) -> dict:
    """
    Import histori harga dari Excel ke tabel price_history.

    Returns:
        dict: { total_rows, imported, skipped, errors, items_unique }
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {filepath}")

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)

    stats = {"total_rows": 0, "imported": 0, "skipped": 0, "errors": 0, "items": set()}
    records = []

    for sheet_name in INVOICE_BB_SHEETS:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' tidak ditemukan di Excel")
            continue

        ws = wb[sheet_name]
        dapur = DAPUR_MAP.get(sheet_name, sheet_name)
        current_date: Optional[date] = None
        header_found = False

        for row in ws.iter_rows(values_only=True):
            # Abaikan baris yang sepenuhnya kosong
            if not row or all(c is None for c in row):
                continue

            # Deteksi baris header tabel (No, Nama Barang, ...)
            if row[0] is not None and str(row[0]).strip() in ("No", "no", "NO"):

                header_found = True
                continue

            if not header_found:
                # Coba ambil tanggal dari baris sebelum header
                if row[0] is not None and is_date_row(row):
                    parsed = parse_tanggal_blok(row[0])
                    if parsed:
                        current_date = parsed
                continue

            # Cek apakah baris tanggal baru (blok berikutnya)
            if row[0] is not None and is_date_row(row):
                parsed = parse_tanggal_blok(row[0])
                if parsed:
                    current_date = parsed
                header_found = False  # next row akan jadi header lagi
                continue

            # Baris data item
            if row[0] is None or not isinstance(row[0], (int, float)):
                continue

            stats["total_rows"] += 1

            nama_raw = row[1]
            if not nama_raw:
                stats["skipped"] += 1
                continue

            nama = normalize_nama(str(nama_raw))
            if not nama:
                stats["skipped"] += 1
                continue

            # Kolom: No(0), NamaBarang(1), Satuan(2), Qty(3),
            #        EstHargaBeli(4), TotalEstimasi(5), HargaBeliNota(6),
            #        TotalBeliNota(7), HargaJual(8), TotalJual(9), ...
            satuan = str(row[2] or "").strip() or None
            qty = parse_rupiah(row[3])

            # Prioritas: Harga Beli Nota (realisasi), fallback ke estimasi
            harga_beli = parse_rupiah(row[6])
            if harga_beli is None:
                harga_beli = parse_rupiah(row[4])

            harga_jual = parse_rupiah(row[8])

            if harga_beli is None and harga_jual is None:
                stats["skipped"] += 1
                continue

            records.append({
                "nama_item": nama,
                "tanggal": current_date or date(2026, 5, 24),
                "harga_beli": harga_beli,
                "harga_jual": harga_jual,
                "qty": qty,
                "satuan": satuan,
                "dapur": dapur,
                "sumber": models.PriceSumber.excel_import,
            })
            stats["items"].add(nama)

    wb.close()

    # Batch insert (hapus data lama dari sumber excel_import dulu)
    try:
        db.query(models.PriceHistory).filter(
            models.PriceHistory.sumber == models.PriceSumber.excel_import
        ).delete()
        db.commit()
    except Exception as e:
        logger.error(f"Gagal hapus data lama: {e}")
        db.rollback()

    # Insert batch
    batch_size = 200
    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        try:
            db.bulk_insert_mappings(models.PriceHistory, batch)
            db.commit()
            stats["imported"] += len(batch)
        except Exception as e:
            logger.error(f"Batch insert error: {e}")
            db.rollback()
            stats["errors"] += len(batch)

    stats["items_unique"] = len(stats.pop("items"))
    logger.info(f"Import selesai: {stats}")
    return stats


def get_histori_harga(nama_item: str, db: Session, limit: int = 90) -> list[dict]:
    """
    Ambil histori harga untuk satu item, terurut dari terlama ke terbaru.
    """
    nama_norm = normalize_nama(nama_item)
    rows = (
        db.query(models.PriceHistory)
        .filter(models.PriceHistory.nama_item == nama_norm)
        .order_by(models.PriceHistory.tanggal.asc())
        .limit(limit)
        .all()
    )
    return [
        {
            "tanggal": str(r.tanggal),
            "harga_beli": float(r.harga_beli) if r.harga_beli else None,
            "harga_jual": float(r.harga_jual) if r.harga_jual else None,
            "qty": float(r.qty) if r.qty else None,
            "satuan": r.satuan,
            "dapur": r.dapur,
        }
        for r in rows
    ]
