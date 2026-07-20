"""Supplier management router - CRUD for vendors/suppliers."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
import models, schemas, auth
from database import get_db
import os, shutil

router = APIRouter()


def generate_kode_supplier(db: Session) -> str:
    count = db.query(func.count(models.Supplier.id)).scalar() + 1
    return f"SUP{count:04d}"


@router.get("/", response_model=list[schemas.SupplierOut])
def list_supplier(
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin,
        models.UserRole.finance, models.UserRole.akuntan
    )),
):
    q = db.query(models.Supplier)
    if is_active is not None:
        q = q.filter(models.Supplier.is_active == is_active)
    return q.order_by(models.Supplier.nama).all()


@router.get("/{supplier_id}", response_model=schemas.SupplierOut)
def get_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.get_current_user),
):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")
    return supplier


@router.post("/", response_model=schemas.SupplierOut)
def create_supplier(
    payload: schemas.SupplierCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin, models.UserRole.finance
    )),
):
    # Check kode unik
    if db.query(models.Supplier).filter(models.Supplier.kode == payload.kode).first():
        raise HTTPException(status_code=400, detail=f"Kode supplier '{payload.kode}' sudah digunakan")

    supplier = models.Supplier(**payload.model_dump())
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    return supplier


@router.put("/{supplier_id}", response_model=schemas.SupplierOut)
def update_supplier(
    supplier_id: int,
    payload: schemas.SupplierUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin, models.UserRole.finance
    )),
):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(supplier, field, value)
    db.commit()
    db.refresh(supplier)
    return supplier


@router.delete("/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin
    )),
):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")

    # Soft delete
    supplier.is_active = False
    db.commit()
    return {"message": "Supplier dinonaktifkan"}


@router.get("/{supplier_id}/hutang-summary")
def supplier_hutang_summary(
    supplier_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin,
        models.UserRole.finance,
    )),
):
    """
    Ringkasan hutang supplier + daftar transaksi belanja yang terkait.
    Digunakan di halaman status pembayaran supplier.
    """
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")

    hutang_list = (
        db.query(models.HutangSupplier)
        .options(joinedload(models.HutangSupplier.pembayaran_list))
        .filter(models.HutangSupplier.supplier_id == supplier_id)
        .order_by(models.HutangSupplier.tanggal.desc())
        .all()
    )

    total_hutang = sum(float(h.jumlah or 0) for h in hutang_list)
    total_terbayar = sum(float(h.jumlah_terbayar or 0) for h in hutang_list)
    total_sisa = sum(float(h.sisa or 0) for h in hutang_list)
    has_hutang = any(h.status != models.HutangStatus.lunas for h in hutang_list)

    return {
        "supplier": {
            "id": supplier.id,
            "nama": supplier.nama,
            "kode": supplier.kode,
            "rekening": supplier.rekening,
            "nama_bank": supplier.nama_bank,
            "kontak": supplier.kontak,
        },
        "has_hutang": has_hutang,
        "total_hutang": total_hutang,
        "total_terbayar": total_terbayar,
        "total_sisa": total_sisa,
        "hutang_list": [
            {
                "id": h.id,
                "nomor_hutang": h.nomor_hutang,
                "tanggal": str(h.tanggal),
                "jatuh_tempo": str(h.jatuh_tempo) if h.jatuh_tempo else None,
                "jumlah": float(h.jumlah or 0),
                "jumlah_terbayar": float(h.jumlah_terbayar or 0),
                "sisa": float(h.sisa or 0),
                "status": h.status.value,
                "deskripsi": h.deskripsi,
                "pembayaran_list": [
                    {
                        "id": p.id,
                        "tanggal_bayar": str(p.tanggal_bayar),
                        "jumlah_bayar": float(p.jumlah_bayar or 0),
                        "metode": p.metode,
                        "referensi": p.referensi,
                        "bukti_bayar_path": p.bukti_bayar_path,
                    }
                    for p in h.pembayaran_list
                ],
            }
            for h in hutang_list
        ],
    }



@router.post("/import-excel")
async def import_supplier_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(auth.require_roles(
        models.UserRole.admin, models.UserRole.super_admin, models.UserRole.finance
    )),
):
    """
    Import data supplier dari file Excel (.xlsx/.xls).
    Kolom yang dikenali: Nama, Kode, Alamat, Kontak/Telepon/HP, Email,
    Kategori, Terms/JatuhTempo, Rekening/No Rekening, Bank/Nama Bank.
    Baris yang sudah ada (berdasarkan nama) akan di-skip.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="Hanya file Excel (.xlsx/.xls) yang diterima")

    import tempfile, os
    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True)

        added = 0
        skipped = 0
        errors = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Cari baris header
            header_row_idx = None
            headers = []
            for i, row in enumerate(rows[:10]):
                row_lower = [str(c).lower().strip() if c else "" for c in row]
                if any("nama" in c or "supplier" in c or "vendor" in c for c in row_lower):
                    header_row_idx = i
                    headers = row_lower
                    break

            if header_row_idx is None:
                continue

            # Map kolom
            col_map = {}
            for j, h in enumerate(headers):
                if "nama" in h and "bank" not in h:
                    col_map.setdefault("nama", j)
                elif "kode" in h:
                    col_map["kode"] = j
                elif "alamat" in h:
                    col_map["alamat"] = j
                elif any(k in h for k in ["kontak", "telp", "telepon", "hp", "phone"]):
                    col_map["kontak"] = j
                elif "email" in h:
                    col_map["email"] = j
                elif "kategori" in h or "jenis" in h:
                    col_map["kategori"] = j
                elif any(k in h for k in ["terms", "jatuh", "tempo"]):
                    col_map["terms_pembayaran"] = j
                elif "rekening" in h or "no rek" in h or "account" in h:
                    col_map["rekening"] = j
                elif "bank" in h:
                    col_map["nama_bank"] = j

            if "nama" not in col_map:
                continue

            # Proses data rows
            for row in rows[header_row_idx + 1:]:
                if not row or not row[col_map["nama"]]:
                    continue

                nama = str(row[col_map["nama"]]).strip()
                if not nama or nama.lower() in ("nama", "supplier", "vendor", "none", ""):
                    continue

                # Cek sudah ada berdasarkan nama
                existing = db.query(models.Supplier).filter(
                    models.Supplier.nama.ilike(nama)
                ).first()
                if existing:
                    skipped += 1
                    continue

                try:
                    kode = str(row[col_map["kode"]]).strip() if "kode" in col_map and row[col_map["kode"]] else generate_kode_supplier(db)
                    # Pastikan kode unik
                    if db.query(models.Supplier).filter(models.Supplier.kode == kode).first():
                        kode = generate_kode_supplier(db)

                    terms = 0
                    if "terms_pembayaran" in col_map and row[col_map["terms_pembayaran"]]:
                        try:
                            terms = int(float(str(row[col_map["terms_pembayaran"]]).strip()))
                        except Exception:
                            terms = 0

                    supplier = models.Supplier(
                        kode=kode,
                        nama=nama,
                        alamat=str(row[col_map["alamat"]]).strip() if "alamat" in col_map and row[col_map["alamat"]] else None,
                        kontak=str(row[col_map["kontak"]]).strip() if "kontak" in col_map and row[col_map["kontak"]] else None,
                        email=str(row[col_map["email"]]).strip() if "email" in col_map and row[col_map["email"]] else None,
                        kategori=str(row[col_map["kategori"]]).strip() if "kategori" in col_map and row[col_map["kategori"]] else None,
                        terms_pembayaran=terms,
                        rekening=str(row[col_map["rekening"]]).strip() if "rekening" in col_map and row[col_map["rekening"]] else None,
                        nama_bank=str(row[col_map["nama_bank"]]).strip() if "nama_bank" in col_map and row[col_map["nama_bank"]] else None,
                        is_active=True,
                    )
                    db.add(supplier)
                    db.flush()
                    added += 1
                except Exception as e:
                    errors.append(f"Baris '{nama}': {str(e)}")

        db.commit()
        return {
            "status": "berhasil",
            "ditambahkan": added,
            "dilewati": skipped,
            "errors": errors[:10],
        }
    except Exception as e:
        raise HTTPException(500, detail=f"Gagal membaca Excel: {str(e)}")
    finally:
        os.unlink(tmp_path)
